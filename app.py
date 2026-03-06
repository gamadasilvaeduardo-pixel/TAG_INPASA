# app.py
# -*- coding: utf-8 -*-

import io
import re
import os
import json
import streamlit as st
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook, Workbook

import gspread
from google.oauth2.service_account import Credentials

from pdf_engine import build_pdf_bytes_mixed

st.set_page_config(page_title="Gerador de Tags", layout="wide")

# =========================
# TIMEZONE (Brasília)
# =========================
BR_TZ = ZoneInfo("America/Sao_Paulo")

def now_str():
    return datetime.now(BR_TZ).strftime("%Y-%m-%d %H:%M:%S")

# =========================
# AUTH
# =========================
ADMIN_USER = "admin"
ADMIN_PASS = "cpcm123"

# senha extra para ações destrutivas/admin
ADMIN_CONFIRM_PASS = "cpcm123"

# =========================
# Persistência simples (JSON local do app)
# =========================
ALLOWLIST_FILE = "users_allowlist.json"
BASE_CACHE_FILE = "bases_cache.json"
PASSWORDS_FILE = "users_passwords.json"

def norm_user(s: str) -> str:
    return (s or "").strip().upper()

def user_password_rule(username: str) -> str:
    # senha padrão inicial
    u = (username or "").strip().lower()
    return f"{u}123"

# =========================
# TAG helpers
# =========================
def norm_tag(s: str) -> str:
    s = (s or "").strip()
    s = s.replace("_", "-")
    s = re.sub(r"\s+", "", s)
    return s.upper()

def get_prefix(tag: str) -> str:
    m = re.match(r"^([A-Z]+)", (tag or "").upper())
    return m.group(1) if m else ""

# =========================
# XLSX helpers (sem pandas)
# =========================
def list_sheets(uploaded_file):
    if uploaded_file is None:
        return []
    bio = io.BytesIO(uploaded_file.getvalue())
    wb = load_workbook(bio, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names

def read_xlsx(uploaded_file, sheet_name=None):
    if uploaded_file is None:
        return [], []
    bio = io.BytesIO(uploaded_file.getvalue())
    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        wb.close()
        return [], []
    headers = [str(h).strip().upper() if h is not None else "" for h in header]
    rows = list(it)
    wb.close()
    return headers, rows

def idx_of(headers_list, *names):
    for n in names:
        n = n.upper()
        if n in headers_list:
            return headers_list.index(n)
    return None

# =========================
# Regra do supervisório
# =========================
def _fix_parentheses(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return s

    cleaned = []
    open_count = 0
    for ch in s:
        if ch == "(":
            open_count += 1
            cleaned.append(ch)
        elif ch == ")":
            if open_count > 0:
                open_count -= 1
                cleaned.append(ch)
        else:
            cleaned.append(ch)

    s2 = "".join(cleaned)
    if open_count > 0:
        s2 += (")" * open_count)
    return s2

def normalize_supervisor_field(text: str) -> str:
    s = (text or "").strip().upper()
    s = _fix_parentheses(s)

    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1].strip()

    s = _fix_parentheses(s)
    return f"({s})" if s else "()"

def clean_prefix_desc(prefix_desc: str) -> str:
    s = (prefix_desc or "").strip().upper()
    s = re.sub(r"\(\s*$", "", s).strip()
    s = re.sub(r"[\-:]\s*$", "", s).strip()
    s = re.sub(r"\s+", " ", s).strip()
    return s

def build_manual_description(prefix_desc: str, supervisor_text: str) -> str:
    p = clean_prefix_desc(prefix_desc)
    sup = normalize_supervisor_field(supervisor_text)
    return f"{p} {sup}".strip() if p else sup

def remove_tag_prefix_inside_parentheses(tag: str, desc_final: str) -> str:
    t = (tag or "").strip().upper()
    d = (desc_final or "").strip().upper()
    pfx = get_prefix(t)
    if not pfx:
        return d
    d = re.sub(rf"\(\s*{re.escape(pfx)}\s*\)", " ", d)
    d = re.sub(r"\s+", " ", d).strip()
    return d

# =========================
# Allowlist + cache bases (JSON)
# =========================
def load_allowlist():
    if os.path.exists(ALLOWLIST_FILE):
        try:
            with open(ALLOWLIST_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                return sorted({norm_user(x) for x in data if str(x).strip()})
        except Exception:
            pass
    return []

def save_allowlist(users):
    users = sorted({norm_user(x) for x in (users or []) if str(x).strip()})
    with open(ALLOWLIST_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)
    return users

def save_bases_cache(base_tags, base_prefix):
    data = {"base_tags": base_tags, "base_prefix": base_prefix, "saved_at": now_str()}
    with open(BASE_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)

def load_bases_cache():
    if os.path.exists(BASE_CACHE_FILE):
        try:
            with open(BASE_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            bt = data.get("base_tags") or {}
            bp = data.get("base_prefix") or {}
            if isinstance(bt, dict) and isinstance(bp, dict):
                return bt, bp, data.get("saved_at", "")
        except Exception:
            pass
    return {}, {}, ""

# =========================
# Passwords (JSON local)
# =========================
def load_passwords():
    if os.path.exists(PASSWORDS_FILE):
        try:
            with open(PASSWORDS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return {norm_user(k): str(v) for k, v in data.items() if str(k).strip()}
        except Exception:
            pass
    return {}

def save_passwords(passwords_dict):
    normed = {}
    for k, v in (passwords_dict or {}).items():
        kk = norm_user(k)
        vv = str(v or "").strip()
        if kk and vv:
            normed[kk] = vv
    with open(PASSWORDS_FILE, "w", encoding="utf-8") as f:
        json.dump(normed, f, ensure_ascii=False, indent=2)
    return normed

def get_user_saved_password(username: str) -> str:
    passwords = st.session_state.get("passwords", {})
    return str(passwords.get(norm_user(username), "")).strip()

def verify_user_password(username: str, password_input: str) -> bool:
    uname = norm_user(username)
    p_in = (password_input or "").strip()

    saved = get_user_saved_password(uname)
    if saved:
        return p_in == saved

    return p_in.lower() == user_password_rule(uname).lower()

def set_user_password(username: str, new_password: str):
    uname = norm_user(username)
    passwords = dict(st.session_state.get("passwords", {}))
    passwords[uname] = (new_password or "").strip()
    passwords = save_passwords(passwords)
    st.session_state["passwords"] = passwords

def reset_user_password_to_default(username: str):
    uname = norm_user(username)
    passwords = dict(st.session_state.get("passwords", {}))
    if uname in passwords:
        del passwords[uname]
    passwords = save_passwords(passwords)
    st.session_state["passwords"] = passwords

# =========================
# Google Sheets LOG
# =========================
def get_gs_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(
        dict(st.secrets["gcp_service_account"]),
        scopes=scopes
    )
    return gspread.authorize(creds)

def get_log_worksheet():
    gc = get_gs_client()
    sheet_id = st.secrets["log_sheet"]["spreadsheet_id"]
    ws_name = st.secrets["log_sheet"].get("worksheet", "LOG")
    sh = gc.open_by_key(sheet_id)

    try:
        ws = sh.worksheet(ws_name)
    except Exception:
        ws = sh.add_worksheet(title=ws_name, rows=2000, cols=10)

    try:
        a1 = ws.acell("A1").value
    except Exception:
        a1 = None

    if (a1 or "").strip().upper() != "DATA_HORA":
        ws.update("A1:F1", [[
            "DATA_HORA", "USUARIO", "TAG", "DESCRICAO_FINAL", "STATUS", "LAYOUT"
        ]])
    return ws

def append_log_rows_gs(rows):
    ws = get_log_worksheet()
    ws.append_rows(rows, value_input_option="USER_ENTERED")

def export_log_xlsx_from_gs() -> bytes:
    ws = get_log_worksheet()
    values = ws.get_all_values()

    wb = Workbook()
    sh = wb.active
    sh.title = "LOG"
    for row in values:
        sh.append(row)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def clear_user_log(username: str) -> int:
    ws = get_log_worksheet()
    values = ws.get_all_values()

    if not values:
        return 0

    header = values[0]
    rows = values[1:]

    user_u = (username or "").strip().upper()

    kept = [header]
    removed = 0
    for r in rows:
        u = (r[1] if len(r) >= 2 else "").strip().upper()
        if u == user_u:
            removed += 1
        else:
            kept.append(r)

    ws.clear()
    ws.update("A1", kept)
    return removed

# =========================
# Session state
# =========================
def ensure_state():
    if "auth" not in st.session_state:
        st.session_state["auth"] = {"logged": False, "role": "", "user": ""}

    if "items_by_user" not in st.session_state:
        st.session_state["items_by_user"] = {}

    if "allowlist" not in st.session_state:
        st.session_state["allowlist"] = load_allowlist()

    if "passwords" not in st.session_state:
        st.session_state["passwords"] = load_passwords()

    if "base_tags" not in st.session_state or "base_prefix" not in st.session_state:
        bt, bp, saved_at = load_bases_cache()
        st.session_state["base_tags"] = bt
        st.session_state["base_prefix"] = bp
        st.session_state["bases_saved_at"] = saved_at

    st.session_state.setdefault("login_user", "")
    st.session_state.setdefault("login_pass", "")

ensure_state()

def get_user_items(user: str):
    m = st.session_state["items_by_user"]
    if user not in m:
        m[user] = []
    return m[user]

def set_user_items(user: str, items):
    st.session_state["items_by_user"][user] = items

# =========================
# UI
# =========================
st.title("Gerador de Tags")

auth = st.session_state["auth"]

# ---------- LOGIN ----------
if not auth["logged"]:
    st.subheader("Login")

    with st.form("login_form", clear_on_submit=False):
        u_in = st.text_input("Usuário", key="login_user", value=st.session_state["login_user"], placeholder="")
        p_in = st.text_input("Senha", key="login_pass", value=st.session_state["login_pass"], type="password", placeholder="")
        do_login = st.form_submit_button("Entrar")

    if do_login:
        u = (u_in or "").strip()
        p = (p_in or "").strip()

        if not u or not p:
            st.error("Informe usuário e senha.")
            st.stop()

        # admin
        if u.lower() == ADMIN_USER and p == ADMIN_PASS:
            st.session_state["auth"] = {"logged": True, "role": "admin", "user": "ADMIN"}
            st.rerun()

        # user
        user_norm = norm_user(u)
        allow = st.session_state["allowlist"]

        if allow and (user_norm not in allow):
            st.error("Usuário não autorizado. Fale com o administrador.")
            st.stop()

        if verify_user_password(user_norm, p):
            st.session_state["auth"] = {"logged": True, "role": "user", "user": user_norm}
            st.rerun()

        st.error("Usuário ou senha inválidos.")

    st.stop()

# ---------- AUTH OK ----------
role = auth["role"]
user = auth["user"]
is_admin = (role == "admin")

top1, top2 = st.columns([3, 1])
with top1:
    st.caption(f"Logado como: **{user}**  | Perfil: **{role.upper()}**")
with top2:
    if st.button("Sair", key="btn_logout"):
        st.session_state["auth"] = {"logged": False, "role": "", "user": ""}
        st.session_state["login_user"] = ""
        st.session_state["login_pass"] = ""
        st.rerun()

# =========================
# PERFIL / SENHA
# =========================
with st.expander("👤 Perfil e senha", expanded=False):
    if is_admin:
        st.info("Perfil ADMIN. A senha do admin continua fixa no código.")
    else:
        st.markdown("### Redefinir minha senha")

        with st.form("form_change_my_password", clear_on_submit=True):
            current_pass = st.text_input("Senha atual", type="password", key="my_current_pass")
            new_pass = st.text_input("Nova senha", type="password", key="my_new_pass")
            confirm_new_pass = st.text_input("Confirmar nova senha", type="password", key="my_confirm_new_pass")
            do_change_pass = st.form_submit_button("Salvar nova senha")

        if do_change_pass:
            if not current_pass.strip() or not new_pass.strip() or not confirm_new_pass.strip():
                st.error("Preencha todos os campos.")
            elif not verify_user_password(user, current_pass):
                st.error("Senha atual inválida.")
            elif len(new_pass.strip()) < 4:
                st.error("A nova senha deve ter pelo menos 4 caracteres.")
            elif new_pass != confirm_new_pass:
                st.error("A confirmação da nova senha não confere.")
            else:
                set_user_password(user, new_pass.strip())
                st.success("Senha alterada com sucesso.")

# =========================
# ADMIN PANEL
# =========================
if is_admin:
    with st.expander("🔒 Admin — Bases + Usuários + LOG", expanded=False):
        st.markdown("### Bases (somente admin)")
        colA, colB = st.columns([1, 1], gap="large")

        with colA:
            up_base = st.file_uploader(
                "Base principal (XLSX) colunas: TAG, DESCRIÇÃO/DESCRICAO",
                type=["xlsx"],
                key="up_base"
            )
            base_sheets = list_sheets(up_base)
            base_sheet = st.selectbox(
                "Aba da Base principal",
                options=base_sheets or ["(envie o arquivo)"],
                index=0,
                key="base_sheet"
            )

        with colB:
            up_prefix = st.file_uploader(
                "Base de prefixos (XLSX) colunas: PREFIXO, PRE_DESCRIÇÃO",
                type=["xlsx"],
                key="up_prefix"
            )
            pref_sheets = list_sheets(up_prefix)
            pref_sheet = st.selectbox(
                "Aba da Base prefixos",
                options=pref_sheets or ["(envie o arquivo)"],
                index=0,
                key="pref_sheet"
            )

        if st.button("Salvar bases", key="btn_save_bases"):
            base_tags = {}
            base_prefix = {}

            headers, rows = read_xlsx(up_base, base_sheet if base_sheets else None)
            if not headers:
                st.error("Envie a base principal.")
                st.stop()

            i_tag = idx_of(headers, "TAG")
            i_desc = idx_of(headers, "DESCRIÇÃO", "DESCRICAO")
            if i_tag is None or i_desc is None:
                st.error("Base principal inválida: precisa de TAG e DESCRIÇÃO/DESCRICAO na linha 1.")
                st.stop()

            for r in rows:
                t = norm_tag(str(r[i_tag]) if i_tag < len(r) and r[i_tag] is not None else "")
                d = str(r[i_desc]).strip() if i_desc < len(r) and r[i_desc] is not None else ""
                if t:
                    base_tags[t] = d.upper().strip()

            headersp, rowsp = read_xlsx(up_prefix, pref_sheet if pref_sheets else None)
            if headersp:
                i_pfx = idx_of(headersp, "PREFIXO")
                i_pd = idx_of(headersp, "PRE_DESCRIÇÃO", "PRE_DESCRICAO", "PRÉ_DESCRIÇÃO", "PRE-DESCRICAO")
                if i_pfx is None or i_pd is None:
                    st.error("Base prefixos inválida: precisa de PREFIXO e PRE_DESCRIÇÃO na linha 1.")
                    st.stop()

                for r in rowsp:
                    pfx = (str(r[i_pfx]).strip().upper() if i_pfx < len(r) and r[i_pfx] is not None else "")
                    pdsc = (str(r[i_pd]).strip() if i_pd < len(r) and r[i_pd] is not None else "")
                    if pfx:
                        base_prefix[pfx] = pdsc.upper().strip()

            st.session_state["base_tags"] = base_tags
            st.session_state["base_prefix"] = base_prefix
            st.session_state["bases_saved_at"] = now_str()
            save_bases_cache(base_tags, base_prefix)
            st.success(f"Bases salvas: {len(base_tags)} TAGs | {len(base_prefix)} prefixos")

        st.caption(
            f"Cache atual: {len(st.session_state['base_tags'])} TAGs | "
            f"{len(st.session_state['base_prefix'])} prefixos | "
            f"salvo em: {st.session_state.get('bases_saved_at') or '—'}"
        )

        st.divider()
        st.markdown("### Usuários permitidos (allowlist)")
        current = st.session_state["allowlist"]
        txt = st.text_area("1 usuário por linha", value="\n".join(current), height=160, key="txt_allowlist")
        if st.button("Salvar allowlist", key="btn_save_allowlist"):
            users = [line.strip() for line in txt.splitlines() if line.strip()]
            st.session_state["allowlist"] = save_allowlist(users)
            st.success(f"Allowlist salva ({len(st.session_state['allowlist'])} usuários).")

        st.divider()
        st.markdown("### Reset de senha de usuário")

        allow_users = st.session_state["allowlist"] or []
        if not allow_users:
            st.info("Não há usuários na allowlist.")
        else:
            reset_user = st.selectbox(
                "Usuário para resetar",
                options=allow_users,
                key="sel_reset_user"
            )

            st.caption(f"O reset remove a senha personalizada e volta para a senha padrão: {reset_user.lower()}123")

            with st.form("form_reset_user_password", clear_on_submit=True):
                admin_reset_pass = st.text_input(
                    "Senha do administrador",
                    type="password",
                    key="admin_reset_pass"
                )
                do_reset = st.form_submit_button("Resetar senha para o padrão")

            if do_reset:
                if not admin_reset_pass.strip():
                    st.error("Informe a senha do administrador.")
                elif admin_reset_pass != ADMIN_CONFIRM_PASS:
                    st.error("Senha do administrador inválida.")
                else:
                    reset_user_password_to_default(reset_user)
                    st.success(f"Senha de {reset_user} resetada para o padrão: {reset_user.lower()}123")

        st.divider()
        st.markdown("### LOG completo (Google Sheets) — somente admin")
        try:
            xbytes = export_log_xlsx_from_gs()
            st.download_button(
                "Baixar LOG completo (XLSX)",
                data=xbytes,
                file_name=f"log_tags_{now_str().replace(':','-').replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_dl_log"
            )
        except Exception as e:
            st.error(f"Falha ao ler LOG do Google Sheets: {e}")

        st.divider()
        st.markdown("### Limpar LOG de usuário (Google Sheets) — somente admin")

        try:
            ws = get_log_worksheet()
            vals = ws.get_all_values()
            users_log = sorted({
                r[1].strip().upper()
                for r in vals[1:]
                if len(r) >= 2 and str(r[1]).strip()
            })

            if not users_log:
                st.info("Não há usuários no LOG para limpar.")
            else:
                user_to_clear = st.selectbox(
                    "Usuário para limpar",
                    options=users_log,
                    key="sel_clear_user"
                )

                st.caption("A limpeza só será executada após confirmação com senha do administrador.")

                with st.form("form_clear_user_log", clear_on_submit=True):
                    confirm_pass = st.text_input(
                        "Senha do administrador",
                        type="password",
                        key="confirm_clear_log_pass"
                    )
                    do_clear = st.form_submit_button("Confirmar limpeza do LOG")

                if do_clear:
                    if not confirm_pass.strip():
                        st.error("Informe a senha do administrador.")
                    elif confirm_pass != ADMIN_CONFIRM_PASS:
                        st.error("Senha inválida. Limpeza cancelada.")
                    else:
                        removed = clear_user_log(user_to_clear)
                        st.success(f"OK — removidos {removed} registros do usuário {user_to_clear}.")

        except Exception as e:
            st.error(f"Falha ao carregar usuários do LOG: {e}")

st.divider()

# =========================
# USER FLOW
# =========================
st.subheader("2) Inserir TAG")

base_tags = st.session_state["base_tags"]
base_prefix = st.session_state["base_prefix"]

tag_raw = st.text_input(
    "TAG",
    value="",
    placeholder="Ex: INC-1608516A / TIT-1800100 / ME-1203019",
    key="inp_tag"
)
tag = norm_tag(tag_raw)

is_square = st.checkbox("TAG QUADRADA (150×150)", value=False, key="chk_square")
layout_name = "square" if is_square else "small"

desc_in = ""
manual_flag = False
base_desc = ""

if tag:
    base_desc = (base_tags.get(tag, "") or "").strip().upper()

    if base_desc:
        st.info("TAG encontrada na base.")
        alter = st.checkbox("Alterar descrição desta TAG", value=False, key=f"alter_{tag}")
        desc_in = st.text_input("Descrição", value=base_desc, disabled=(not alter), key=f"desc_{tag}")
        desc_in = (desc_in or "").strip().upper()
        manual_flag = False
    else:
        st.warning("TAG NÃO encontrada na base.")
        prefix = get_prefix(tag)
        pre_desc = base_prefix.get(prefix, "")
        st.caption(f"Sugestão por prefixo ({prefix}): {pre_desc or '—'}")

        sup = st.text_input(
            "Descrição do supervisório / IO List",
            value="",
            placeholder="Ex: BOMBA DE TESTE (P-410) ou (BOMBA DE TESTE (P-410",
            key=f"sup_{tag}"
        )

        desc_final = build_manual_description(pre_desc, sup)
        desc_in = remove_tag_prefix_inside_parentheses(tag, desc_final)
        manual_flag = True

if st.button("Adicionar à lista", key="btn_add"):
    if not tag:
        st.error("Informe a TAG.")
    elif not desc_in.strip():
        st.error("Informe a descrição.")
    else:
        items = get_user_items(user)
        ts = now_str()

        changed = bool(base_desc and (desc_in.strip().upper() != base_desc.strip().upper()))

        updated = False
        for it in items:
            if it["tag"] == tag:
                it.update({
                    "desc": desc_in.strip().upper(),
                    "layout": layout_name,
                    "manual": bool(manual_flag),
                    "changed": bool(changed),
                    "base_desc": base_desc,
                    "updated_at": ts,
                    "updated_by": user,
                })
                updated = True
                break

        if not updated:
            items.append({
                "tag": tag,
                "desc": desc_in.strip().upper(),
                "layout": layout_name,
                "manual": bool(manual_flag),
                "changed": bool(changed),
                "base_desc": base_desc,
                "created_at": ts,
                "created_by": user,
                "updated_at": ts,
                "updated_by": user,
            })

        set_user_items(user, items)
        st.success("Adicionado/atualizado.")

st.divider()

st.subheader("3) Lista atual (sua lista)")
items = get_user_items(user)

if items:
    st.dataframe(
        [
            {
                "TAG": it["tag"],
                "DESCRIÇÃO": it["desc"],
                "LAYOUT": "150×150" if it.get("layout") == "square" else "100×50",
                "MANUAL": "SIM" if it.get("manual") else "",
                "ALTERADO": "SIM" if it.get("changed") else "",
            }
            for it in items
        ],
        width="stretch",
        height=360
    )
else:
    st.write("Nenhuma TAG adicionada (na sua lista).")

c1, c2 = st.columns([1, 1])

with c1:
    if st.button("Limpar lista", key="btn_clear_list"):
        set_user_items(user, [])
        st.success("Sua lista foi limpa.")

with c2:
    if items:
        ts_pdf = now_str()
        rows_log = []
        for it in items:
            status = "MANUAL" if it.get("manual") else ("ALTERADO" if it.get("changed") else "BASE")
            layout = "150×150" if it.get("layout") == "square" else "100×50"
            rows_log.append([ts_pdf, user, it["tag"], it["desc"], status, layout])

        try:
            pdf_bytes = build_pdf_bytes_mixed([
                (it["tag"], it["desc"], it.get("layout", "small"))
                for it in items
            ])
        except Exception as e:
            st.error(f"Falha ao gerar PDF: {e}")
            st.stop()

        def _on_download():
            append_log_rows_gs(rows_log)

        st.download_button(
            "Gerar/baixar PDF",
            data=pdf_bytes,
            file_name="tags.pdf",
            mime="application/pdf",
            key="btn_dl_pdf",
            on_click=_on_download,
        )
    else:
        st.button("Gerar/baixar PDF", disabled=True, key="btn_dl_pdf_disabled")
